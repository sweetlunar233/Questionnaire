import random
from django.shortcuts import render
from django.http import JsonResponse

# Create your views here.
from .models import User,Survey
from .models import BaseQuestion,BlankQuestion,ChoiceQuestion,ChoiceOption,RatingQuestion
from .models import Answer,BlankAnswer,ChoiceAnswer,RatingAnswer
from .models import Submission,SurveyStatistic,Template,RewardOffering,UserRewardRecord   

import json
from django.core.mail import BadHeaderError, send_mail
from django.http import HttpResponse, HttpResponseRedirect

from django.core.mail import EmailMessage

from itsdangerous import URLSafeTimedSerializer as utsr
import base64
from django.conf import settings as django_settings
from django.utils import timezone
from django.db import transaction 

from rest_framework.views import APIView
import itertools

from itertools import chain  
from operator import attrgetter 

serveAddress="http:127.0.0.1:8080"

#普通问卷的展示界面：
def display_answer_normal(request,username,questionnaireId,submissionId):
    user=User.objects.get(username=username)
    if user is None:
        return HttpResponse(content='User not found', status=404) 
        
    survey=Survey.objects.get(SurveyID=questionnaireId)
    if survey is None:
        return HttpResponse(content='Questionnaire not found', status=404)   
    
    submission=Submission.objects.get(SubmissionID=submissionId)
    if submission is None:
        return HttpResponse(content='Submission not found', status=404)  
    
    all_questionList_iterator = itertools.chain(BlankQuestion.objects.filter(Survey=survey).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','CorrectAnswer','QuestionNumber','QuestionID').all(),
                                                    ChoiceQuestion.objects.filter(Survey=survey).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','OptionCnt','QuestionNumber','QuestionID').all(),
                                                    RatingQuestion.objects.filter(Survey=survey).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','QuestionID','QuestionNumber').all())
                                                    
    # 将迭代器转换为列表 (按QuestionNumber递增排序)
    all_questions_list = list(all_questionList_iterator)
    all_questions_list.sort(key=lambda x: x['QuestionNumber']) 

    #print(all_questions_list.length())
    questionList=[]
    #print(all_questions)
    for question in all_questions_list:
        if question["Category"]==1 or question["Category"]==2:    #选择题

            #该单选题的用户选项:当前问卷当前submission(如果用户未选，则找不到对应的答案记录)
            if question["Category"]==1:
                optionAnswer_query=ChoiceAnswer.objects.filter(Submission=submission,Question=question["QuestionID"])  #只有一条记录
                
                #用户未填该单选题
                if not optionAnswer_query.exists():
                    answer=-1
                #用户填了这个单选题，有一条答案记录
                else:
                    answer=optionAnswer_query.first().ChoiceOptions.OptionID
            
            #该多选题的用户选项:当前问卷当前submission
            else:
                optionAnswer_query=ChoiceAnswer.objects.filter(Submission=submission,Question=question["QuestionID"])#一或多条记录
                #用户未填该多选题
                if not optionAnswer_query.exists():answer=[]
                #用户填了这个多选题，有一条/多条答案记录
                else:
                    answer=[]
                    for optionAnswer in optionAnswer_query:
                        answer.append(optionAnswer.ChoiceOptions.OptionID)

            optionList=[]
            #所有选项
            options_query=ChoiceOption.objects.filter(Question=question["QuestionID"])
            for option in options_query:
                optionList.append({'content':option.Text,'optionNumber':option.OptionNumber,'isCorrect':option.IsCorrect,'optionId':option.OptionID})
            questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                    'isNecessary':question["IsRequired"],'score':question["Score"],'optionCnt':question["OptionCnt"],
                                    'optionList':optionList,'Answer':answer})
            
        elif question["Category"]==3:                  #填空题
            #该填空题的用户答案:有且仅有一条记录
            blankAnswer_query=BlankAnswer.objects.filter(Submission=submission,Question=question["QuestionID"])
            #用户未填该填空题
            if not blankAnswer_query.exists():
                answer=""
            else:
                answer=blankAnswer_query.first().Content
            
            questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                    'isNecessary':question["IsRequired"],'score':question["Score"],
                                    'correctAnswer':question["CorrectAnswer"],'Answer':answer})

        elif question["Category"]==4:                  #评分题
            #该评分题的用户答案:有且仅有一条记录
            ratingAnswer_query=RatingAnswer.objects.filter(Submission=submission,Question=question["QuestionID"])
            #用户未填该评分题
            if not ratingAnswer_query.exists():
                answer=0
            else:
                #print("123")
                answer=ratingAnswer_query.first().Rate

            questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                    'isNecessary':question["IsRequired"],'score':question["Score"],'Answer':answer})

    data={'Title':survey.Title,'description':survey.Description,'questionList':questionList}
    return JsonResponse(data)


#考试问卷的展示界面：
def display_answer_test(request,username,questionnaireId,submissionId):
    print("start display_answer_test")
    # print(submissionId)
    user=User.objects.get(username=username)
    if user is None:
        return HttpResponse(content='User not found', status=404) 
        
    survey=Survey.objects.get(SurveyID=questionnaireId)
    if survey is None:
        return HttpResponse(content='Questionnaire not found', status=404)   
    
    submission=Submission.objects.get(SubmissionID=submissionId)
    if submission is None:
        return HttpResponse(content='Submission not found', status=404)  
    score=submission.Score
    
    all_questionList_iterator = itertools.chain(BlankQuestion.objects.filter(Survey=survey).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','CorrectAnswer','QuestionNumber','QuestionID').all(),
                                                    ChoiceQuestion.objects.filter(Survey=survey).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','OptionCnt','QuestionNumber','QuestionID').all(),
                                                    RatingQuestion.objects.filter(Survey=survey).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','QuestionID','QuestionNumber').all())
    
    # 将迭代器转换为列表 (按QuestionNumber递增排序)
    all_questions_list = list(all_questionList_iterator)
    all_questions_list.sort(key=lambda x: x['QuestionNumber']) 

    questionList=[]
    #print(all_questions)
    for question in all_questions_list:
        if question["Category"]==1 or question["Category"]==2:    #选择题

            #该单选题的用户选项:当前问卷当前submission(如果用户未选，则找不到对应的答案记录)
            if question["Category"]==1:
                optionAnswer_query=ChoiceAnswer.objects.filter(Submission=submission,Question=question["QuestionID"])  #只有一条记录
                
                #用户未填该单选题
                if not optionAnswer_query.exists():
                    answer=-1
                #用户填了这个单选题，有一条答案记录
                else:
                    answer=optionAnswer_query.first().ChoiceOptions.OptionID
            
            #该多选题的用户选项:当前问卷当前submission
            else:
                optionAnswer_query=ChoiceAnswer.objects.filter(Submission=submission,Question=question["QuestionID"])#一或多条记录
                #用户未填该多选题
                if not optionAnswer_query.exists():answer=[]
                #用户填了这个多选题，有一条/多条答案记录
                else:
                    answer=[]
                    for optionAnswer in optionAnswer_query:
                        answer.append(optionAnswer.ChoiceOptions.OptionID)

            optionList=[]
            #所有选项
            options_query=ChoiceOption.objects.filter(Question=question["QuestionID"])
            for option in options_query:
                optionList.append({'content':option.Text,'optionNumber':option.OptionNumber,'isCorrect':option.IsCorrect,'optionId':option.OptionID})
            questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                    'isNecessary':question["IsRequired"],'score':question["Score"],'optionCnt':question["OptionCnt"],
                                    'optionList':optionList,'Answer':answer})
            
        elif question["Category"]==3:                  #填空题
            #该填空题的用户答案:有且仅有一条记录
            blankAnswer_query=BlankAnswer.objects.filter(Submission=submission,Question=question["QuestionID"])
            #用户未填该填空题
            if not blankAnswer_query.exists():
                answer=""
            else:
                answer=blankAnswer_query.first().Content
            
            questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                    'isNecessary':question["IsRequired"],'score':question["Score"],
                                    'correctAnswer':question["CorrectAnswer"],'Answer':answer})

        elif question["Category"]==4:                  #评分题
            #该评分题的用户答案:有且仅有一条记录
            ratingAnswer_query=RatingAnswer.objects.filter(Submission=submission,Question=question["QuestionID"])
            #用户未填该评分题
            if not ratingAnswer_query.exists():
                answer=0
            else:
                answer=ratingAnswer_query.first().Rate

            questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                    'isNecessary':question["IsRequired"],'score':question["Score"],'Answer':answer})


    data={'Title':survey.Title,'description':survey.Description,'questionList':questionList,'score':score}
    # print(questionList[0])
    return JsonResponse(data)


#问卷填写界面：向前端传输问卷当前暂存的填写记录
class GetStoreFillView(APIView):
    def get(self, request, *args, **kwargs):  
        # 从查询参数中获取userName和surveyID   
        userName = kwargs.get('userName')  
        surveyID = kwargs.get('surveyID')   
        submissionID=kwargs.get('submissionID')  
        
        user=User.objects.get(username=userName)
        if user is None:
            return HttpResponse(content='User not found', status=404) 
        
        survey=Survey.objects.get(SurveyID=surveyID)
        if survey is None:
            return HttpResponse(content='Questionnaire not found', status=404) 
          
        
        #从问卷广场界面进入：查找该用户是否有该问卷未提交的填写记录
        if submissionID=="-1":
            submission_query=Submission.objects.filter(Respondent=user,Survey=survey,Status='Unsubmitted')
            if submission_query.exists():
                submissionID=submission_query.first().SubmissionID  #找到未填写的记录
                duration=submission_query.first().Interval
                submission = submission_query.first()
                # newsubmissionID = submissionID
            
            else:      #不存在：创建一条新的填写记录
                submission=Submission.objects.create(Survey=survey,Respondent=user,Status="Unsubmitted",
                                                    Interval=0)
                duration=0
                submissionID=submission.SubmissionID
                # newsubmissionID = submission.SubmissionID
                # return HttpResponse(content='Submission not existed', status=404) 

        #submissionID=-2时,只传回问卷题干(同问卷编辑的GET接口)
        elif submissionID=="-2":
            print("--here---")
            all_questionList_iterator = itertools.chain(BlankQuestion.objects.filter(Survey=survey).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','CorrectAnswer','QuestionNumber','QuestionID').all(),
                                                    ChoiceQuestion.objects.filter(Survey=survey).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','OptionCnt','QuestionNumber','QuestionID').all(),
                                                    RatingQuestion.objects.filter(Survey=survey).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','QuestionNumber','QuestionID').all())
                                                    
            # 将迭代器转换为列表  
            all_questions_list = list(all_questionList_iterator)
            all_questions_list.sort(key=lambda x: x['QuestionNumber']) 

            questionList=[]

            for question in all_questions_list:
                if question["Category"]==1 or question["Category"]==2:    #选择题
                    optionList=[]
                    #将所有选项顺序排列
                    options_query=ChoiceOption.objects.filter(Question=question["QuestionID"]).order_by('OptionNumber')
                    for option in options_query:
                        optionList.append({'content':option.Text,'optionNumber':option.OptionNumber,'isCorrect':option.IsCorrect,
                                       'optionID':option.OptionID,'MaxSelectablePeople':option.MaxSelectablePeople})
                    questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                     'isNecessary':question["IsRequired"],'score':question["Score"],'optionCnt':question["OptionCnt"],
                                     'optionList':optionList})
                
                elif question["Category"]==3:                  #填空题
                
                    questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                     'isNecessary':question["IsRequired"],'score':question["Score"],'correctAnswer':question["CorrectAnswer"]})

                elif question["Category"]==4:                  #评分题
                    questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                     'isNecessary':question["IsRequired"],'score':question["Score"]})

        
            data={'Title':survey.Title,'category':survey.Category,'TimeLimit':survey.TimeLimit,
                'description':survey.Description,'questionList':questionList}
            return JsonResponse(data)
        
        submission=Submission.objects.filter(SubmissionID=submissionID).first()
        print(submission.Interval)
        # print(submission)
        if not submission:
            return HttpResponse(content='Submission not found', status=404) 
    
        Title=survey.Title
        Description=survey.Description
        category=survey.Category
        TimeLimit=survey.TimeLimit
        #people=survey.QuotaLimit
        
        '''1.以下部分与问卷编辑界面的get函数类似，拿到题干'''
        '''2.拿到当前submissionID对应填写记录'''
        all_questionList_iterator = itertools.chain(BlankQuestion.objects.filter(Survey=survey).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','CorrectAnswer','QuestionNumber','QuestionID').all(),
                                                    ChoiceQuestion.objects.filter(Survey=survey).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','OptionCnt','QuestionNumber','QuestionID','MaxSelectable').all(),
                                                    RatingQuestion.objects.filter(Survey=survey).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','QuestionID','QuestionNumber').all())
                                                    
        all_questions_list = list(all_questionList_iterator)

        # 将迭代器转换为列表 (按QuestionNumber递增排序)--顺序展示
        if survey.IsOrder:
            all_questions_list.sort(key=lambda x: x['QuestionNumber']) 
        
        #print(all_questions_list.length())
        questionList=[]
        #print(all_questions)
        for question in all_questions_list:
            if question["Category"]==1 or question["Category"]==2:    #选择题
                #print(question['MaxSelectable'])
                print(question)
                print(question['OptionCnt'])

                #该单选题的用户选项:当前问卷当前submission(如果用户未选，则找不到对应的答案记录)
                if question["Category"]==1:
                    optionAnswer_query=ChoiceAnswer.objects.filter(Submission=submission,Question=question["QuestionID"])  #只有一条记录
                    #用户未填该单选题
                    if not optionAnswer_query.exists():
                        answer=-1
                    #用户填了这个单选题，有一条答案记录
                    else:
                        answer=optionAnswer_query.first().ChoiceOptions.OptionID
                
                #该多选题的用户选项:当前问卷当前submission
                else:
                    optionAnswer_query=ChoiceAnswer.objects.filter(Submission=submission,Question=question["QuestionID"])#一或多条记录
                    #用户未填该多选题
                    if not optionAnswer_query.exists():answer=[]
                    #用户填了这个多选题，有一条/多条答案记录
                    else:
                        answer=[]
                        for optionAnswer in optionAnswer_query:
                            answer.append(optionAnswer.ChoiceOptions.OptionID)

                optionList=[]
                #将所有选项顺序排列
                print("***")
                options_query=ChoiceOption.objects.filter(Question=question["QuestionID"]).order_by('OptionNumber')
                for option in options_query:
                    optionList.append({'content':option.Text,'optionNumber':option.OptionNumber,'isCorrect':option.IsCorrect,
                                       'optionId':option.OptionID,'MaxSelectablePeople':option.MaxSelectablePeople})
                
                if survey.Category == 3 and survey.IsOrder == False: #选项乱序展示
                    random.shuffle(optionList)
                
                questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                     'isNecessary':question["IsRequired"],'score':question["Score"],'optionCnt':question["OptionCnt"],
                                     'optionList':optionList,'Answer':answer,'max':question['MaxSelectable']})
                
            elif question["Category"]==3:                  #填空题
                #该填空题的用户答案:有且仅有一条记录
                blankAnswer_query=BlankAnswer.objects.filter(Submission=submission,Question=question["QuestionID"])
                #用户未填该填空题
                if not blankAnswer_query.exists():
                    answer=""
                else:
                    answer=blankAnswer_query.first().Content
                
                questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                     'isNecessary':question["IsRequired"],'score':question["Score"],
                                     'correctAnswer':question["CorrectAnswer"],'Answer':answer})

            elif question["Category"]==4:                  #评分题
                
                #该评分题的用户答案:有且仅有一条记录
                ratingAnswer_query=RatingAnswer.objects.filter(Submission=submission,Question=question["QuestionID"])
                #用户未填该评分题
                if not ratingAnswer_query.exists():
                    answer=0
                else:
                    answer=ratingAnswer_query.first().Rate

                questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                     'isNecessary':question["IsRequired"],'score':question["Score"],'Answer':answer})
        
        #题干乱序展示
        if survey.Category == 3 and survey.IsOrder == False:
            random.shuffle(questionList)

        #传回题干和填写记录
        data={'Title':survey.Title,'category':survey.Category,'TimeLimit':survey.TimeLimit,
            'description':survey.Description,'questionList':questionList,'duration':submission.Interval, 'submissionID':submissionID}
        
        return JsonResponse(data)
        

#问卷填写界面：从前端接收用户的填写记录(POST)
def get_submission(request):
    if(request.method=='POST'):
        try:
            # print("start get_submission")
            body=json.loads(request.body)
            surveyID=body['surveyID']    #问卷id
            status=body['status']  #填写记录状态
            submissionID=body['submissionID']   #填写记录ID
            username=body['username']     #填写者
            submissionList=body['question']     #填写记录
            duration=body['duration']  
            score=body['score'] 

            # 新加的
            publishDate=body['date']  #日期

            # print("lorian")
            # print(submissionID)

            # print(submissionList)

            survey=Survey.objects.get(SurveyID=surveyID)
            if survey is None:
                return HttpResponse(content='Questionnaire not found',status=404)
            
            user=User.objects.get(username=username)
            if user is None:
                return HttpResponse(content='User not found',status=404)

            #当前不存在该填写记录，创建：  //实际上用不到，在getStoreFill的时候就给不存在的submission创建新的Id了
            if submissionID==-1:
                submission=Submission.objects.create(Survey=survey,Respondent=user,
                                             SubmissionTime=timezone.now(),Status=status,
                                             Interval=duration,Score=score)
                print(submission.SubmissionTime)
            
            #已存在，删除填写记录的所有内容
            else:
                submission=Submission.objects.get(SubmissionID=submissionID)
                if submission is None:
                    return HttpResponse(content='Submission not found',status=404)
                submission.Score=score
                submission.Status=status
                submission.Interval=duration
                submission.SubmissionTime=timezone.now()    #更新为当前时间
                submission.save()
                
                #所有选择题的填写记录
                ChoiceAnswer_query=ChoiceAnswer.objects.filter(Submission=submission)
                if ChoiceAnswer_query.exists():
                    for choiceAnswer in ChoiceAnswer_query:
                        choiceAnswer.delete()
                
                #所有填空题的填写记录
                BlankAnswer_query=BlankAnswer.objects.filter(Submission=submission)
                if BlankAnswer_query.exists():
                    for blankAnswer in BlankAnswer_query:
                        blankAnswer.delete()
                
                #所有评分题的填写记录
                RatingAnswer_query=RatingAnswer.objects.filter(Submission=submission)
                if RatingAnswer_query.exists():
                    for ratingAnswer in RatingAnswer_query:
                        ratingAnswer.delete()

            # 新加的
            survey.PublishDate=publishDate
            survey.save()

            for submissionItem in submissionList:
                # print("TieZhu")
                questionID=submissionItem["questionID"]     #问题ID
                answer=submissionItem['value']        #用户填写的答案
                category=submissionItem['category']     #问题类型（用于后续区分，解决不同种类问题的QuestionID会重复的问题）

                #print(category)
                #question = BaseQuestion.objects.get(QuestionID=questionID).select_subclasses()   #联合查询

                '''
                question_iterator=itertools.chain(ChoiceQuestion.objects.filter(QuestionID=questionID),
                                                    BlankQuestion.objects.filter(QuestionID=questionID),
                                                    RatingQuestion.objects.filter(QuestionID=questionID))
                question_list=list(question_iterator)
                question=question_list[0]
                print(question)
                print(question_list)
                # print(question["Category"])
                # print(question.Category)'''

                questionNewList=[]
                choiceQuestion_query=ChoiceQuestion.objects.filter(QuestionID=questionID,Category=category)
                if choiceQuestion_query.exists():
                    questionNewList.append(choiceQuestion_query.first())

                blankQuestion_query=BlankQuestion.objects.filter(QuestionID=questionID,Category=category)
                if blankQuestion_query.exists():
                    questionNewList.append(blankQuestion_query.first())

                ratingQuestion_query=RatingQuestion.objects.filter(QuestionID=questionID,Category=category)
                if ratingQuestion_query.exists():
                    questionNewList.append(ratingQuestion_query.first())
                
                question=questionNewList[0]
                
                # print("123154654")

                # print(question.CorrectAnswer)
                if question is None:
                    return HttpResponse(content='Question not found',status=404)

                if question.Category==1:     #单选题：Answer为选项ID
                    if answer==-1: continue       #返回-1，代表用户没填该单选题
                    option=ChoiceOption.objects.get(OptionID=answer)     #用户选择的选项
                    if option is None:
                        return HttpResponse(content="Option not found",status=404)
                    choiceAnswer=ChoiceAnswer.objects.create(Question=question,Submission=submission,ChoiceOptions=option)
                    choiceAnswer.save()

                    #若已提交，报名问卷的必填选择题中，选择的对应选项人数-1
                    if status=='Submitted' and survey.Category==2 and question.IsRequired==True:
                        print(option.MaxSelectablePeople)

                        if option.MaxSelectablePeople<=0:
                            data={'message':False,'content':'报名人数已满'}
                            return JsonResponse(data)
                        
                        else:
                            option.MaxSelectablePeople-=1
                            option.save()


                elif question.Category==2:     #多选题：Answer为选项ID的数组
                    #为每个用户选择的选项，创建一条ChoiceAnswer记录
                    for optionID in answer:
                        option=ChoiceOption.objects.get(OptionID=optionID)     #用户选择的选项
                        if option is None:
                            return HttpResponse(content="Option not found",status=404)
                        choiceAnswer=ChoiceAnswer.objects.create(Question=question,Submission=submission,ChoiceOptions=option)
                        choiceAnswer.save()

                        #若已提交，报名问卷的必填选择题中，选择的对应选项人数-1
                        if status=='Submitted' and survey.Category==2 and question.IsRequired==True:
                            if option.MaxSelectablePeople<=0:
                                data={'message':False,'content':'报名人数已满'}
                                return JsonResponse(data)
                            else:
                                option.MaxSelectablePeople-=1
                                option.save()

                elif question.Category==3:     #填空题：answer为填写的内容
                    blankAnswer=BlankAnswer.objects.create(Question=question,Submission=submission,Content=answer)
                    blankAnswer.save()
                
                elif question.Category==4:      #评分题：answer为填写的内容
                    # print(answer)
                    ratingAnswer=RatingAnswer.objects.create(Question=question,Submission=submission,Rate=answer)
                    ratingAnswer.save()

            user.zhibi+=50
            user.save()
                
        except json.JSONDecodeError:  
            return JsonResponse({'error': 'Invalid JSON body'}, status=400)
        except Exception as e:  
            return JsonResponse({'error': str(e)}, status=500) 
    data={'message':True,'submissionId':submissionID}
    # print(submissionID)
    return JsonResponse(data)
    #return JsonResponse({'error': 'Invalid request method'}, status=405)


#问卷编辑界面：向前端传输问卷设计内容
class GetQuestionnaireView(APIView):
    def get(self, request, survey_id, *args, **kwargs):  
        design = request.GET.get('design', 'false')  # 默认为'false'  
        design = design.lower() == 'true'  # 将字符串转换为布尔值  
        survey=Survey.objects.get(SurveyID=survey_id)
        if survey is None:
            return HttpResponse(content='Questionnaire not found', status=400) 
        title=survey.Title
        catecory=survey.Category
        #people=survey.QuotaLimit
        TimeLimit=survey.TimeLimit

        '''
        blank_questions = list(BlankQuestion.objects.filter(Survey=survey).values_list('id', 'QuestionNumber'))  
        choice_questions = list(ChoiceQuestion.objects.filter(Survey=survey).values_list('id', 'QuestionNumber'))  
        rating_questions = list(RatingQuestion.objects.filter(Survey=survey).values_list('id', 'QuestionNumber'))  

        # 将这些列表合并，并基于QuestionNumber进行排序  
        combined_questions = sorted(chain(blank_questions, choice_questions, rating_questions), key=lambda x: x[1])
        '''

        all_questionList_iterator = itertools.chain(BlankQuestion.objects.filter(Survey=survey).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','CorrectAnswer','QuestionNumber','QuestionID').all(),
                                                    ChoiceQuestion.objects.filter(Survey=survey).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','OptionCnt','QuestionNumber','MaxSelectable','QuestionID').all(),
                                                    RatingQuestion.objects.filter(Survey=survey).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','QuestionNumber','QuestionID').all())
                                                    
        # 将迭代器转换为列表  
        all_questions_list = list(all_questionList_iterator)
        all_questions_list.sort(key=lambda x: x['QuestionNumber']) 

        questionList=[]

        #print(all_questions)
        for question in all_questions_list:
            if question["Category"]==1 or question["Category"]==2:    #选择题
                optionList=[]
                #将所有选项顺序排列
                options_query=ChoiceOption.objects.filter(Question=question["QuestionID"]).order_by('OptionNumber')
                for option in options_query:
                    optionList.append({'content':option.Text,'optionNumber':option.OptionNumber,'isCorrect':option.IsCorrect,
                                       'optionID':option.OptionID,'MaxSelectablePeople':option.MaxSelectablePeople})
                questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                     'isNecessary':question["IsRequired"],'score':question["Score"],'optionCnt':question["OptionCnt"],
                                     'optionList':optionList,'max':question['MaxSelectable']})
                
            elif question["Category"]==3:                  #填空题
                
                questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                     'isNecessary':question["IsRequired"],'score':question["Score"],'correctAnswer':question["CorrectAnswer"]})

            elif question["Category"]==4:                  #评分题
                questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                     'isNecessary':question["IsRequired"],'score':question["Score"]})

        
        data={'Title':survey.Title,'category':survey.Category,'TimeLimit':survey.TimeLimit,
              'description':survey.Description,'questionList':questionList}
        
        return JsonResponse(data, status=200)


#问卷编辑界面：从前端接收问卷的设计内容
def save_qs_design(request):
    if(request.method=='POST'):
        try:
            body=json.loads(request.body)
            surveyID=body['surveyID']    #问卷id
            title=body['title']  #问卷标题
            catecory=body['category']   #问卷类型（普通0、投票1、报名2、考试3）
            isOrder=body['isOrder'] #是否顺序展示（考试问卷）
            #people=body['people']   #报名人数（报名问卷）
            timelimit=body['timeLimit']
            username=body['userName']   #创建者用户名
            description=body['description'] #问卷描述
            Is_released=body['Is_released'] #保存/发布

            questionList=body['questionList']   #问卷题目列表

            # 新加的
            publishDate=body['date'] #日期

            print(questionList)
            user=User.objects.get(username=username)
            if user is None:        
                return HttpResponse(content='User not found', status=400) 
            
            #当前不存在该问卷，创建：
            if surveyID==-1:
                survey=Survey.objects.create(Owner=user,Title=title,
                                             Description=description,Is_released=Is_released,
                                             Is_open=True,Is_deleted=False,Category=catecory,
                                             TotalScore=0,TimeLimit=timelimit,IsOrder=isOrder
                                            )
                print("TieZhu")
                #survey.QuotaLimit=people
            #已有该问卷的编辑记录
            else:
                survey=Survey.objects.get(SurveyID=surveyID)
                if survey is None:
                    return HttpResponse(content='Questionnaire not found', status=400) 
                
                survey.Title=title
                survey.Is_released=Is_released
                survey.Description=description
                survey.Category=catecory
                survey.TimeLimit=timelimit
                survey.IsOrder=isOrder
                #survey.QuotaLimit=people    #该问卷的报名人数
                survey.save()

                #该问卷的所有选择题
                choiceQuestion_query=ChoiceQuestion.objects.filter(Survey=survey)
                for choiceQuestion in choiceQuestion_query:
                    #删除该选择题的所有选项
                    choiceOption_query=ChoiceOption.objects.filter(Question=choiceQuestion)
                    for choiceOption in choiceOption_query:
                        choiceOption.delete()
                    choiceQuestion.delete()

                #删除该问卷的所有填空题
                blankQuestion_query=BlankQuestion.objects.filter(Survey=survey)
                for blankQuestion in blankQuestion_query:
                    blankQuestion.delete()
                
                #删除该问卷的所有评分题
                ratingQuestion_query=RatingQuestion.objects.filter(Survey=survey)
                for ratingQuestion in ratingQuestion_query:
                    ratingQuestion.delete()

            # 新加的
            survey.PublishDate=publishDate
            survey.save()

            index=1
            for question in questionList:
                if question["type"]==1 or question["type"]==2:        #单选/多选

                    optionList=question['optionList']
                    
                    question=ChoiceQuestion.objects.create(Survey=survey,Text=question["question"],IsRequired=question["isNecessary"],
                                                                QuestionNumber=index,Score=question["score"],Category=question["type"],
                                                                OptionCnt=question["optionCnt"],MaxSelectable=question['max'])
                    question.save()

                    #所有选项:
                    jdex=1
                    for option in optionList:
                        option=ChoiceOption.objects.create(Question=question,Text=option["content"],IsCorrect=option["isCorrect"],
                                                           OptionNumber=jdex,MaxSelectablePeople=option['MaxSelectablePeople'])
                        option.save()
                        jdex=jdex+1

                
                elif question["type"]==3:                          #填空
                    # print(question)
                    question=BlankQuestion.objects.create(Survey=survey,Text=question["question"],IsRequired=question["isNecessary"],
                                                        Score=question["score"],QuestionNumber=index,Category=question["type"],
                                                            CorrectAnswer=question["correctAnswer"])
                    question.save()  
                
                else:                                           #评分题
                    question=RatingQuestion.objects.create(Survey=survey,Text=question["question"],IsRequired=question["isNecessary"],
                                                              Score=question["score"],QuestionNumber=index,Category=question["type"])
                    question.save()
                index=index+1
            return HttpResponse(content='Questionnaire saved successfully', status=200) 
        except json.JSONDecodeError:  
            return JsonResponse({'error': 'Invalid JSON body'}, status=400)
        except Exception as e:  
            return JsonResponse({'error': str(e)}, status=500) 
    return JsonResponse({'error': 'Invalid request method'}, status=405)


#填写记录
def delete_filled_qs(request):
    if(request.method=='POST'):
        try:
            body=json.loads(request.body)
            submissionID=body
            submission=Submission.objects.get(SubmissionID=submissionID)     #对应填写记录
            if submission is None:
                return JsonResponse({'error': 'No ID provided'}, status=400) 
            submission.delete()

        except json.JSONDecodeError:  
            return JsonResponse({'error': 'Invalid JSON body'}, status=400)
        except Exception as e:  
            return JsonResponse({'error': str(e)}, status=500) 
    data = {"message": "True"}
    return JsonResponse(data)

def update_or_delete_released_qs(request):
    if(request.method=='POST'):
        try:
            body=json.loads(request.body)
            flag=body['flag']

        #创建者删除已发布的问卷(将问卷状态改为Is_deleted=True)
        #所有该问卷填写者处，该问卷的状态修改为已删除；填写者刷新问卷管理界面，保留被删除项，但无法继续填写
            if flag==1:
                qsID=body['id']
                if qsID is None:
                    return JsonResponse({'error': 'No ID provided'}, status=400) 
                qs=Survey.objects.filter(SurveyID=qsID).first()     #对应问卷
                qs.Is_deleted=True
                qs.Is_released=False
                qs.save()

                submission_query=Submission.objects.filter(Survey=qs)   #该问卷的所有填写记录
            
                # 使用 for 循环遍历 submission_query  
                with transaction.atomic():  # 你可以使用事务确保操作的原子性  
                    for submission in submission_query:  
                        #该填写已提交：状态不变
                        #该填写未提交：填写状态改为'Deleted'(已被创建者删除)
                        if submission.Status=='Unsubmitted':
                            submission.Status='Deleted'
                            submission.save()
                
            
            #更新发布状态
            else:
                qsID=body['id']
                if qsID is None:
                    return JsonResponse({'error': 'No ID provided'}, status=400) 
                qs=Survey.objects.filter(SurveyID=qsID).first()     #对应问卷

                #当前未发布，改为发布状态：
                if qs.Is_open==False:
                    qs.Is_open=True
                
                #当前已发布，撤回
                else:
                    qs.Is_open=False
                qs.save()

        except json.JSONDecodeError:  
            return JsonResponse({'error': 'Invalid JSON body'}, status=400)
        except Exception as e:  
            return JsonResponse({'error': str(e)}, status=500) 
    data={"message":"True"}
    return JsonResponse(data)
    #return JsonResponse({'error': 'Invalid request method'}, status=405)


#删除未发布的问卷(直接从数据库移除)
def delete_unreleased_qs(request):
    if(request.method=='POST'):
        try:
            body=json.loads(request.body)
            qsID=body
            if qsID is None:
                return JsonResponse({'error': 'No ID provided'}, status=400) 
            qs=Survey.objects.filter(SurveyID=qsID).first()
            if qs is None:  
                return JsonResponse({'error': 'No questionnaire found with the given ID'}, status=404)
            qs.delete()

            data={'message':'True'}
            return JsonResponse(data)
        except json.JSONDecodeError:  
            return JsonResponse({'error': 'Invalid JSON body'}, status=400)
        except Exception as e:  
            return JsonResponse({'error': str(e)}, status=500) 
    return JsonResponse({'error': 'Invalid request method'}, status=405)

#当前用户已创建未发布的问卷
def get_drafted_qs(request,username):
    if(request.method=='GET'):
        user=User.objects.get(username=username)
        qs_query=Survey.objects.filter(Owner=user,Is_released=False)
        data_list=[{'Title':survey.Title,'PublishDate':survey.PublishDate,'SurveyID':survey.SurveyID,'Category':survey.Category} for survey in qs_query]
        data={'data':data_list}
        return JsonResponse(data)
    return JsonResponse({'error': 'Invalid request method'}, status=405)

#当前用户发布的问卷
def get_released_qs(request,username):
    if(request.method=='GET'):
        user=User.objects.get(username=username)
        qs_query=Survey.objects.filter(Owner=user,Is_released=True,Is_deleted=False)    #不显示已删除问卷

        data_list=[]
        for survey in qs_query:
            submissionCnt=Submission.objects.filter(Survey=survey,Status__in=['Submitted','Graded']).count()  #该问卷已提交的填写份数
            data_list.append({'Title':survey.Title,'PublishDate':survey.PublishDate,'SurveyID':survey.SurveyID,
                    'Category':survey.Category,'Description':survey.Description,'FilledPeople':submissionCnt, 'IsOpening':survey.Is_open})
        data={'data':data_list}
        return JsonResponse(data)
    return JsonResponse({'error': 'Invalid request method'}, status=405)

#当前用户的填写记录(包括被创建者删除的问卷的填写记录)
def get_filled_qs(request,username):
    if(request.method=='GET'):
        user=User.objects.get(username=username)
        submission_query=Submission.objects.filter(Respondent=user)
        data_list=[]

         # 使用 for 循环遍历 submission_query  
        with transaction.atomic():  # 你可以使用事务确保操作的原子性  
            for submission in submission_query:
                status=submission.Status
                if status=="Unsubmitted":
                    status_Chinese="未提交"
                elif status=="Submitted" or status=="Graded":
                    status_Chinese="已提交"
                else:
                    status_Chinese="已删除"
                # 新加的
                data_list.append({'Title':submission.Survey.Title,'PublishDate':submission.SubmissionTime,
                                  'SurveyID':submission.Survey.SurveyID,'Category':submission.Survey.Category,
                                  'Description':submission.Survey.Description,'Status':status_Chinese,
                                  'SubmissionID':submission.SubmissionID})
        data={'data':data_list}
        return JsonResponse(data)
    return JsonResponse({'error': 'Invalid request method'}, status=405)

#问卷管理界面：进入填写时，检查当前问卷的Is_open状态；若为False，则创建者已暂停收集，不可再填写
def check_qs_open_stautus(request,questionnaireId):
    qs=Survey.objects.get(SurveyID=questionnaireId)
    if qs is None:
        return HttpResponse(content="Questionnaire not found",status=404)
    if qs.Is_open==False:
        data={"message":False,"content":"该问卷已暂停收集"}
        return JsonResponse(data)
    else:
        data={"message":True,"content":"可开始/修改填写"}

#问卷广场：检查投票/考试问卷
def check_qs(request,username,questionnaireId,type):
    user=User.objects.get(username=username)
    if user is None:
        return HttpResponse(content="User not found",status=404)
    qs=Survey.objects.get(SurveyID=questionnaireId)
    if qs is None:
        return HttpResponse(content="Questionnaire not found",status=404)
    
    #投票问卷:每个用户只可提交一次
    if qs.Category==1:
        submission_query=Submission.objects.filter(Respondent=user,Survey=qs)
        if submission_query.exists():
            submission=submission_query.first()
            if submission.Status=='Unsubmitted':
                data={'message':False,"content":"对于当前问卷，您有未提交的填写记录"}
            elif submission.Status=='Submitted':
                data={'message':False,"content":"您完成投票，不可重复投票"}
            else:
                data={'message':False,"content":"当前问卷已被撤回"}
        else:
            data={'message':"True","content":"可以开始/继续填写"}
        return JsonResponse(data)
    
    #考试问卷：每个用户只可提交一次
    elif qs.Category==3:
        submission_query=Submission.objects.filter(Respondent=user,Survey=qs)
        if submission_query.exists():
            submission=submission_query.first()
            if submission.Status=='Unsubmitted':
                data={'message':False,"content":"对于当前问卷，您有未提交的填写记录"}
            elif submission.Status=='Graded':
                data={'message':False,"content":"您已完成当前考试"}
            else:
                data={'message':False,"content":"当前问卷已被撤回"}
        else:
            data={'message':"True","content":"可以开始/继续填写"}
        return JsonResponse(data)
    
    #报名问卷：超过人数，不可以再报名
    elif qs.Category==2:
        print("TieZhu")
        #检查是否超人数(检查每个必填选择题的所有选项，是否都超人数)
        submission_query=Submission.objects.filter(Respondent=user,Survey=qs)

        choiceQuestion_query=ChoiceQuestion.objects.filter(Survey=qs,Category__in=[1,2],IsRequired=True)
        if choiceQuestion_query.exists():
            #每个必填选项
            for choiceQuestion in choiceQuestion_query:
                isFull=True
                choiceOption_query=ChoiceOption.objects.filter(Question=choiceQuestion)
                #每个选项的剩余人数
                for choiceOption in choiceOption_query:
                    print(choiceOption.MaxSelectablePeople)
                    if choiceOption.MaxSelectablePeople>0:
                        isFull=False
                
                if isFull==True:
                    data={'message':False,"content":"当前报名人数已满"}
                    return JsonResponse(data)

        '''
        currentCnt=Submission.objects.filter(Respondent=user,Survey=qs).count()

        if currentCnt>=qs.QuotaLimit:
            data={'message':False,"content":"当前报名人数已满"}
            return JsonResponse(data)
        '''

        #检查是否有未提交的填写记录
        unsubmitted_query=Submission.objects.filter(Respondent=user,Survey=qs,Status="Unsubmitted")
        if unsubmitted_query.exists():
            data={'message':False,"content":"对于当前问卷，您有未提交的填写记录"}
        
        data={'message':"True","content":"可以开始/继续填写"}
        return JsonResponse(data)   

    #普通问卷
    else: 
        #检查是否有未提交的填写记录
        unsubmitted_query=Submission.objects.filter(Respondent=user,Survey=qs,Status="Unsubmitted")
        if unsubmitted_query.exists():
            data={'message':False,"content":"对于当前问卷，您有未提交的填写记录"}
        else:
            data={'message':"True","content":"可以开始/继续填写"}

        return JsonResponse(data)   
    
#问卷广场：所有问卷
def get_all_released_qs(request):
    if(request.method=='GET'):
        qs_query=Survey.objects.filter(Is_released=True,Is_open=True).order_by("-PublishDate")
        data_list=[]

        for survey in qs_query:
            reward=RewardOffering.objects.filter(Survey=survey).first()
            if reward is not None:
                data_list.append({'Title':survey.Title,'PostMan':survey.Owner.username,'PublishDate':survey.PublishDate,
                                  'SurveyID':survey.SurveyID,'categoryId':survey.Category,'Description':survey.Description,
                                  'Reward':reward.Zhibi,'HeadCount':reward.AvailableQuota})
            else:
                data_list.append({'Title':survey.Title,'PostMan':survey.Owner.username,'PublishDate':survey.PublishDate,
                                  'SurveyID':survey.SurveyID,'categoryId':survey.Category,'Description':survey.Description,
                                  'Reward':None})
        data={'data':data_list}
        return JsonResponse(data)
    return JsonResponse({'error': 'Invalid request method'}, status=405)


'''个人中心界面'''
#购买商店中的头像
def modify_photo_in_shop(request):
    if(request.method=='POST'):
        try:
            body=json.loads(request.body)
            username=body['username']
            user=User.objects.get(username=username)
            if user is None:
                return JsonResponse({'error': 'No user found'}, status=400) 
            
            photonumber = body['photonumber']
            status = body['status']
            #修改头像
            photonumber = body['photonumber']
            status = body['status']
            user.set_array_element(photonumber,status)

            #修改纸币
            zhibi=body['money']
            user.zhibi=zhibi
            user.save()
            
            photos_data = json.loads(user.own_photos)  
            data={'ownphotos':photos_data}
            return JsonResponse(data)

        except json.JSONDecodeError:  
            return JsonResponse({'error': 'Invalid JSON body'}, status=400)
        except Exception as e:  
            return JsonResponse({'error': str(e)}, status=500) 
    return JsonResponse({'error': 'Invalid request method'}, status=405)

#获取个人信息
def get_user_info(request,username):
    if(request.method=='GET'):
        try:
            user=User.objects.get(username=username)
            if user is None:
                return JsonResponse({'error': 'No user found'}, status=400) 
            
            photo=user.get_used_element()
            data={'password':user.password,'email':user.email,'zhibi':user.zhibi,'photo':photo}
            return JsonResponse(data)
        except json.JSONDecodeError:  
            return JsonResponse({'error': 'Invalid JSON body'}, status=400)
        except Exception as e:  
            return JsonResponse({'error': str(e)}, status=500) 
    return JsonResponse({'error': 'Invalid request method'}, status=405)

#修改个人信息
def modify_user_info(request):
    if(request.method=='POST'):
        try:
            body=json.loads(request.body)
            username=body['username']
            flag=body['flag']
            user=User.objects.get(username=username)
            if user is None:
                return JsonResponse({'error': 'No user found'}, status=400) 

            #修改除头像外的其他信息
            if flag==1:
                email=body['email']
                password=body['password']
                user.email=email
                user.password=password
                user.save()
            
            #修改头像：
            elif flag==2:
                photonumber = body['photonumber']
                status = body['status']
                user.set_array_element(photonumber,status)
                user.save()
            
            else:
                # 参数不正确或缺失  
                return JsonResponse({'error': 'Invalid or missing parameters'}, status=400)

        except json.JSONDecodeError:  
            return JsonResponse({'error': 'Invalid JSON body'}, status=400)
        except Exception as e:  
            return JsonResponse({'error': str(e)}, status=500) 
    data={"message":"True"}
    return JsonResponse(data)


class Token:
    def __init__(self, security_key):
        self.security_key = security_key
        # salt是秘钥的编码
        self.salt = base64.encodebytes(security_key.encode('utf-8'))
        #security_key是settings.py中SECURITY_KEY
        #salt是经过base64加密的SECURITY_KEY

    # 生成token,token中可以保存一段信息，这里我们选择保存username
    def generate_validate_token(self, username):
        serializer = utsr(self.security_key)            #生成令牌serializer
        return serializer.dumps(username, self.salt)    #username在令牌中被编码
        #将带有token的验证链接发送至注册邮箱

    # 验证token
    def confirm_validate_token(self, token, expiration=3600):
        serializer = utsr(self.security_key)
        return serializer.loads(token, salt=self.salt, max_age=expiration)

    # 删除token
    def remove_validate_token(self, token):
        serializer = utsr(self.security_key)
        return serializer.loads(token, salt=self.salt)

token_confirm = Token(django_settings.SECRET_KEY)
def get_token(request):

    url = serveAddress+'user/' + token_confirm.generate_validate_token(username='username')
    '''此处将这个url发送到客户邮箱，我们这里就不进行邮件发送的操作了'''
    return HttpResponse(status=200,content=True)

def send_registration_email(request):
    if(request.method=='POST'):
        body=json.loads(request.body)
        username=body['username']
        password=body['password']
        email=body['email']

        # print(username)
        # print(email)

        if(email==False):
            # print("!")
            user_queryset=User.objects.filter(username=username)
            user=user_queryset.first()
            #return HttpResponse(status=200,content=username)
            if not user_queryset.exists():
                data={'message':"1"}
                return JsonResponse(data)
                #return HttpResponse(status=200,content="1")
            elif(password!=user.password):
                data={'message':"2"}
                return JsonResponse(data)
                #return HttpResponse(status=200,content="2")
            else:
                photos_data = json.loads(user.own_photos)  
                data={
                    'message':"0",
                    'username':user.username,
                    'password':user.password,
                    'email':user.email,
                    'ownphotos':photos_data,
                    'zhibi':user.zhibi,
                }
            return JsonResponse(data)

        user1=User.objects.filter(username=username)
        if user1.exists():
            # print("!")
            data={
                'message': "same username",
            }
            return JsonResponse(data)
        
        user2=User.objects.filter(email=email)
        if user2.exists():
            # print("!!")
            data={
                'message': "same email",
            }
            return JsonResponse(data)

        #创建新用户(尚未邮箱验证,非有效用户)
        user=User.objects.create(username=username,email=email,
                                     password=password,CreateDate=timezone.now(),isActive=False)
        user.save()

        #生成令牌
        token = token_confirm.generate_validate_token(username)
        #active_key = base64.encodestring(userName)
        url="/login"

        #发送邮件
        subject="'纸翼传问'新用户注册"
        message=("Hello,"+username+"! 欢迎注册“纸翼传问”!\n"
                     +"请点击以下链接，以激活新账户:\n"
                     +serveAddress+url+token)

        email=EmailMessage(subject=subject,body=message,from_email="1658441344@qq.com",
                            to=[email],reply_to=["1658441344@qq.com"])
        #email.attach_file('/images/weather_map.png')
        email.send()

        data={
            'message': "register success",
        }
        return JsonResponse(data)
    return HttpResponse(status=200,content=True)

#用户点击邮箱链接,调用视图activate_user(),验证激活用户:
def activate_user(request,token):
    try:username=token_confirm.confirm_validate_token(token)
    except:
        return HttpResponse("抱歉，验证链接已过期，请重新注册。")
    try:user=User.objects.get(username=username)
    except User.DoesNotExist:
        return HttpResponse("抱歉，当前用户不存在，请重新注册。")
    user.is_active=True
    user.save()
    return HttpResponse(status=200,content=True)

#额外需要的包
import pandas as pd
from io import BytesIO
import openpyxl

#交叉分析
def cross_analysis(request, QuestionID1, QuestionID2):
    if request.method == 'GET':
        
        question1 = ChoiceQuestion.objects.get(QuestionID=QuestionID1)
        question2 = ChoiceQuestion.objects.get(QuestionID=QuestionID2)

        if QuestionID1 is None or QuestionID2 is None:
            return JsonResponse({'error': 'Missing QuestionID(s)'}, status=404)
        
        if question1.Survey.SurveyID != question2.Survey.SurveyID:
            return JsonResponse({'error': 'Two questions are not from the same questionnaire.'}, status=404)
        
        if question1.Category!=1 and question1.Category!=2:
            print("3")
            return JsonResponse({'error':'Question1 is not a choice question.'},status=404)
        
        if question2.Category!=1 and question2.Category!=2:
            print("4")
            return JsonResponse({'error':'Question2 is not a choice question.'},status=404)
        

        #问题1的所有选项
        all_options_iterator_1=itertools.chain(ChoiceOption.objects.filter(Question=question1).values('OptionID','OptionNumber','Text').all())
                                  
        all_options_list_1 = list(all_options_iterator_1)
        all_options_list_1.sort(key=lambda x: x['OptionNumber']) 

        #问题2的所有选项
        all_options_iterator_2=itertools.chain(ChoiceOption.objects.filter(Question=question2).values('OptionID','OptionNumber','Text').all())
                                  
        all_options_list_2 = list(all_options_iterator_2)
        all_options_list_2.sort(key=lambda x: x['OptionNumber']) 


        crossCount=[]      #数组：选每个交叉选项的人数
        crossText=[]       #数组：每个交叉选项的内容
        for option1 in all_options_list_1:
            for option2 in all_options_list_2:
                crossText.append('-'.join([option1['Text'],option2['Text']]))
                print("莫问题")
                # crossCount.append('-'.join([str(ChoiceAnswer.objects.filter(Question=question1,ChoiceOptions=option1['OptionID']).count()),
                #                             str(ChoiceAnswer.objects.filter(Question=question2,ChoiceOptions=option2['OptionID']).count())]))
                submission_option1 = ChoiceAnswer.objects.filter(Question=question1,ChoiceOptions=option1['OptionID']).values_list('Submission', flat=True).distinct()
                cnt = ChoiceAnswer.objects.filter(Submission__in=submission_option1,Question=question2,ChoiceOptions=option2['OptionID']).count()
                crossCount.append(cnt)
        print(crossCount)
        print(crossText)
        data={'crossCount':crossCount,'crossText':crossText}
        
        return JsonResponse(data)

#下载表格
from openpyxl import Workbook
import datetime
from urllib.parse import quote  

def download_submissions(request, surveyID):
    if request.method == 'GET':
        try:
            print("-here")
            survey = Survey.objects.get(SurveyID=surveyID)
            if survey is None:
                return HttpResponse(content='Questionniare not found.', status=404)  
            print(survey)

            wb = Workbook()
            wb.encoding='utf-8'
            ws=wb.active    #获取第一个工作表


            #第一行内容
            content=['填写者','提交答卷时间','所用时间']
            ws.cell(1,1).value='填写者'
            ws.cell(1,2).value='提交答卷时间'
            ws.cell(1,3).value='所用时间'

            #该问卷的所有问题
            all_questionList_iterator = itertools.chain(BlankQuestion.objects.filter(Survey=survey).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','CorrectAnswer','QuestionNumber','QuestionID').all(),
                                                        ChoiceQuestion.objects.filter(Survey=survey).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','OptionCnt','QuestionNumber','QuestionID').all(),
                                                        RatingQuestion.objects.filter(Survey=survey).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','QuestionNumber','QuestionID').all())
                                                    
            # 将迭代器转换为列表  
            all_questions_list = list(all_questionList_iterator)
            all_questions_list.sort(key=lambda x: x['QuestionNumber']) 

            for i in range(0,len(all_questions_list)):
                question=all_questions_list[i]
                ws.cell(1,i+4,'、'.join([str(i+1),question['Text']])).value='、'.join([str(i+1),question['Text']])    #题号+题干

            #找出该问卷的所有已提交/已打分的填写记录，按提交时间从前向后排序
            submission_list=list(Submission.objects.filter(Survey=survey, Status__in=['Submitted', 'Graded']).values(
                'SubmissionID','SubmissionTime','Respondent','Survey','Interval','Score'
            ))
            if len(submission_list)==0:
                return HttpResponse(content='No submission records available.',status=404)
        
            submission_list.sort(key=lambda x: x['SubmissionTime'])   

            sub_index=1
            ques_index=3
            for submission in submission_list:
                sub_index+=1
                user=User.objects.get(UserID=submission['Respondent'])
                ws.cell(sub_index,1).value=user.username

                ws.cell(sub_index,2).value=submissionTime=submission['SubmissionTime'].strftime('%Y-%m-%d %H:%M:%S')
            
                if survey.Category==3:  #考试问卷：所用时间为Duration；其他问卷：'--'
                    ws.cell(sub_index,3).value=submission['Interval']
                else:
                    ws.cell(sub_index,3).value='--'
            
                ques_index=3
                for question in all_questions_list:
                    ques_index+=1

                    print(question['Category'])
                    if question['Category']==1 or question['Category']==2:
                        #该题的所有选项
                        all_options_answer_list=list(ChoiceAnswer.objects.filter(Question=question['QuestionID'],Submission=submission['SubmissionID'])
                                                      .values('ChoiceOptions'))

                        if len(all_options_answer_list)==0:   #没有选择选项
                            ws.cell(sub_index,ques_index).value=' '

                        else: 
                            optionNumberList=[]
                            for answer in all_options_answer_list:
                                print('***')
                                option=ChoiceOption.objects.filter(OptionID=answer['ChoiceOptions']).first()
                                optionNumberList.append(str(option.OptionNumber))
                            optionNumberList.sort()     #选项顺序升序排列
                            print(optionNumberList)

                            ws.cell(sub_index,ques_index).value=','.join(optionNumberList)
                    
                    elif question['Category']==3:      #填空题
                        answer=BlankAnswer.objects.get(Question=question['QuestionID'],Submission=submission['SubmissionID'])
                        if answer is None:   
                            ws.cell(sub_index,ques_index).value=' '
                        else: 
                            ws.cell(sub_index,ques_index).value=answer.Content

                    else:           #评分题
                        answer=RatingAnswer.objects.get(Question=question['QuestionID'],Submission=submission['SubmissionID'])
                        if answer is None:   
                            ws.cell(sub_index,ques_index).value=' '
                        else: 
                            ws.cell(sub_index,ques_index).value=answer.Rate

                if survey.Category==3:      #考试问卷：增加该填写记录的总得分
                    ws.cell(sub_index,ques_index).value=submission['Score']

            print('finished')
            # 准备写入到IO中
            output = BytesIO()
            #ws.save()
            wb.save(output)     #Excel文件内容保存到IO中

            output.seek(0)	 # 重新定位到开始

            # 设置HttpResponse的类型
            response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
            formatted_now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            file_name = '-'.join([survey.Title, '填写记录统计']) + '_' + formatted_now + '.xlsx' 
            print(file_name)

            file_name = quote(file_name)	 # 使用urlquote()方法解决中文无法使用的问题
            response['Content-Disposition'] = 'attachment; filename=%s' % file_name
            print(response)
            return response

        except Exception as e:
            print(e)
            return JsonResponse({'message': '导出表格失败!'})

    return JsonResponse({'error': 'Invalid request method'}, status=405)

from django.db.models import Count, Sum, Q

#数据分析：向前端传输该问卷的所有题目及填写内容的统计数据
def survey_statistics(request, surveyID):
    if (request.method=='GET'):
        #问卷
        survey = Survey.objects.filter(SurveyID=surveyID)
        if not survey.exists():
            return JsonResponse({'error': 'Survey not found'}, status=404)


        print("start survey_statistics")
        
        survey = Survey.objects.filter(SurveyID=surveyID).first()
        print(survey.Title)
        # print("lorian")
        
        
        all_questionList_iterator = itertools.chain(BlankQuestion.objects.filter(Survey=survey).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','CorrectAnswer','QuestionNumber','QuestionID').all(),
                                                    ChoiceQuestion.objects.filter(Survey=survey).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','OptionCnt','QuestionNumber','QuestionID','MaxSelectable').all(),
                                                    RatingQuestion.objects.filter(Survey=survey).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','QuestionID','QuestionNumber').all())

        # 将迭代器转换为列表 (按QuestionNumber递增排序)                                        
        all_questions_list = list(all_questionList_iterator)
        all_questions_list.sort(key=lambda x: x['QuestionNumber']) 

        #print(all_questions_list.length())
        questionList=[]
        #print(all_questions)
        print("lorian1")
        for question in all_questions_list:
            if question["Category"]==1 or question["Category"]==2:    #选择题
                # print("选择题")
                # print(question)
                 
                #该问题的所有选项
                # all_options_iterator=itertools.chain(ChoiceOption.objects.filter(Question=question).values('OptionNumber','Text'))               
                # all_options_list = list(all_options_iterator)

                all_options_list = list(ChoiceOption.objects.filter(Question=question['QuestionID']).values('OptionID','OptionNumber','Text'))
                all_options_list.sort(key=lambda x: x['OptionNumber']) 

                optionCount=[]      #数组：选每个选项的人数
                optionText=[]       #数组：每个选项的内容
                for option in all_options_list:
                    optionText.append(option['Text'])
                    optionCount.append(ChoiceAnswer.objects.filter(Question=question['QuestionID'],ChoiceOptions=option['OptionID']).count())

                questionList.append({'Content':question["Text"],'Text':optionText,'Count':optionCount,'type':question['Category'],"questionId":question['QuestionID']})

                print("lorian2")
                
                    
            elif question["Category"]==3:       #填空题
                answerText=[]   #数组：填空的内容
                answerCount=[]  #数组：填该内容的人数


                print("www")
                #text_list=list(BlankAnswer.objects.filter(Question=question['QuestionID']).values('Content'))

                text_counts = BlankAnswer.objects.filter(Question=question['QuestionID']).values('Content').annotate(count=Count('Content'))
                for item in text_counts:
                    answerText.append(item['Content'])
                    answerCount.append(item['count'])

                questionList.append({'Content':question["Text"],'Text':answerText,'Count':answerCount,'type':question['Category'],"questionId":question['QuestionID']})

                print("lorian3")
            
            else:       #评分题
                answerText=[]   #数组：评分的分数
                answerCount=[]  #数组：该评分的人数

                rate_counts=RatingAnswer.objects.filter(Question=question['QuestionID']).values('Rate').annotate(count=Count('Rate'))
                for item in rate_counts:
                    answerText.append(item['Rate'])
                    answerCount.append(item['count'])

                questionList.append({'Content':question["Text"],'Text':answerText,'Count':answerCount,'type':question['Category'],"questionId":question['QuestionID']})

                print("lorian4")
                
        print(questionList)

        #传回题干和填写记录的统计数据
        data={'title':survey.Title,'category':survey.Category,'TimeLimit':survey.TimeLimit,
            'description':survey.Description,'questionList':questionList}
        return JsonResponse(data)
    